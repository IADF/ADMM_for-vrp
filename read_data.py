import include
import openpyxl
import csv


##################################
#   read_data()
#   input:void
#   output:
#   用途:将表格中的数据实例化为对象
#   by:noir
##################################
def read_data():
    #   文件读出
    xlsx_file = openpyxl.load_workbook('input_node_100.xlsx')
    sheet_file = xlsx_file.get_sheet_by_name("Sheet1")

    ###############################node#################################
    #   node[id]
    #   出发仓库节点数据
    include.node.append(include.Node())
    include.node[0].node_id = 0
    include.node[0].type = 1
    include.node[0].longitude = 116.571614
    include.node[0].latitude = 39.792844
    include.node[0].pack_total_weight = 0
    include.node[0].pack_total_volume = 0
    include.node[0].first_receive_time = 0
    include.node[0].last_receive_time = 960

    #   客户节点数据
    for i in range(2, include.customer_size + 1):
        include.node.append(include.Node())
        include.node[i - 1].node_id = sheet_file.cell(column=1, row=i).value
        include.node[i - 1].type = sheet_file.cell(column=2, row=i).value
        include.node[i - 1].longitude = sheet_file.cell(column=3, row=i).value
        include.node[i - 1].latitude = sheet_file.cell(column=4, row=i).value
        include.node[i - 1].pack_total_weight = sheet_file.cell(column=5, row=i).value
        include.node[i - 1].pack_total_volume = sheet_file.cell(column=6, row=i).value
        include.node[i - 1].first_receive_time = sheet_file.cell(column=7, row=i).value
        include.node[i - 1].last_receive_time = sheet_file.cell(column=8, row=i).value

    #   终点仓库节点数据
    include.node.append(include.Node())
    include.node[100].node_id = 100
    include.node[100].type = 1
    include.node[100].longitude = 0
    include.node[100].latitude = 0
    include.node[100].pack_total_weight = 0
    include.node[100].pack_total_volume = 0
    include.node[100].first_receive_time = 0
    include.node[100].last_receive_time = 960
    #################################link###########################################
    #   link[from][to]
    #   读取文件
    with open('input_link_100.csv', 'rt') as f:
        csv_file = csv.reader(f)
        i = -1  # 临时变量（用于生成link对象）
        #   将文件数据写入对象
        for row in csv_file:
            if row[0] == "ID":
                continue
            if int(row[1]) != i:
                i += 1
                include.link.append([])
            if int(row[2]) != len(include.link[i]):
                include.link[i].append(include.Link())
            if int(row[4]) > include.transport_time_limit:  # 运输时长过长的线路直接省去
                include.link[i].append(include.Link())
                continue

            include.link[i].append(include.Link())
            print(int(row[0]))
            include.link[i][int(row[2])].link_id = int(row[0])
            include.link[i][int(row[2])].from_node = int(row[1])
            include.link[i][int(row[2])].to_node = int(row[2])
            include.link[i][int(row[2])].distance = int(row[3])
            include.link[i][int(row[2])].spend_time = int(row[4])

            if int(row[0]) == 9999:
                break
    print(dir(include.link[0][100]))

    #########################Agent##############################
    #   agent[vehicle_id]
    for i in range(include.vehicle_fleet_size):
        include.agent.append(include.Agent())
        include.agent[i].agent_id = i
        include.agent[i].from_node_id = include.begin_node_id
        include.agent[i].to_node_id = include.end_node_id
        include.agent[i].agent_volume = 0
        include.agent[i].agent_weight = 0
        include.agent[i].running_time = 0
        include.agent[i].running_distance = 0
        include.agent[i].departure_time = 0
        include.agent[i].arrival = 0

    print("read_date complete")
    return 0
